from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .calculations import parse_money


BLUE = "1F5F9F"
BLUE_SOFT = "E7F0FA"
GREEN = "DCEFD7"
RED = "F7D9D7"
GRAY = "EEF2F7"
INK = "17212F"
MUTED = "66758A"
MONEY_FMT = '"$"#,##0'


def build_excel_workbook(snapshot: dict[str, Any], report: dict[str, Any] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _summary_sheet(ws, snapshot, report or {})
    _sacs_sheet(wb.create_sheet("SACS"), snapshot)
    _tcc_sheet(wb.create_sheet("TCC Totals"), snapshot)
    _accounts_sheet(wb.create_sheet("Accounts"), snapshot)
    _trust_sheet(wb.create_sheet("Trust Assets"), snapshot)
    _liabilities_sheet(wb.create_sheet("Liabilities"), snapshot)
    _inputs_sheet(wb.create_sheet("Inputs"), snapshot)
    for sheet in wb.worksheets:
        _finish_sheet(sheet)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _summary_sheet(ws, snapshot: dict[str, Any], report: dict[str, Any]) -> None:
    _title(ws, "AW Quarterly Report Workbook", 1, 1, 4)
    rows = [
        ("Client", snapshot["client"].get("household_name", "")),
        ("Report Date", snapshot.get("report_date", "")),
        ("Report ID", report.get("id", "")),
        ("Generated At", report.get("created_at", "")),
        ("Status", report.get("status", "generated")),
        ("Export Status", report.get("export_status", "not_configured")),
    ]
    _key_values(ws, 3, rows)
    totals = snapshot.get("totals", {})
    total_rows = [
        ("Monthly Inflow", totals.get("inflow")),
        ("Monthly Outflow", totals.get("outflow")),
        ("Monthly Excess", totals.get("excess")),
        ("Private Reserve Target", totals.get("private_reserve_target")),
        ("Client 1 Retirement", totals.get("client_1_retirement")),
        ("Client 2 Retirement", totals.get("client_2_retirement")),
        ("Non-Retirement", totals.get("non_retirement")),
        ("Trust", totals.get("trust")),
        ("Grand Total", totals.get("grand_total")),
        ("Liabilities Separate", totals.get("liabilities")),
    ]
    _section(ws, "Calculated Totals", 11, 1)
    _key_values(ws, 12, total_rows, money=True)


def _sacs_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "SACS Cashflow", 1, 1, 4)
    totals = snapshot.get("totals", {})
    rows = [
        ("Inflow", totals.get("inflow")),
        ("Outflow", totals.get("outflow")),
        ("Excess", totals.get("excess")),
        ("Private Reserve Balance", totals.get("private_reserve_balance")),
        ("Investment Account Balance", totals.get("investment_account_balance")),
        ("Insurance Deductibles", totals.get("insurance_deductibles")),
        ("Private Reserve Target", totals.get("private_reserve_target")),
    ]
    _key_values(ws, 3, rows, money=True)
    _section(ws, "Deductible Detail", 12, 1)
    _table(ws, 13, ["Label", "Amount", "Source"], [[d.get("label", ""), _money(d.get("amount")), d.get("source", "Profile")] for d in snapshot.get("deductible_items", [])], money_cols={2})


def _tcc_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "TCC Net Worth Totals", 1, 1, 4)
    totals = snapshot.get("totals", {})
    rows = [
        ("Client 1 Retirement Total", totals.get("client_1_retirement")),
        ("Client 2 Retirement Total", totals.get("client_2_retirement")),
        ("Non-Retirement Total", totals.get("non_retirement")),
        ("Trust Total", totals.get("trust")),
        ("Grand Total", totals.get("grand_total")),
        ("Liabilities Total - Separate", totals.get("liabilities")),
    ]
    _key_values(ws, 3, rows, money=True)


def _accounts_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "Accounts", 1, 1, 9)
    rows = []
    for a in snapshot.get("accounts", []):
        rows.append([
            a.get("source", "Profile"),
            _owner(a.get("owner_index")),
            a.get("category", "").replace("_", " ").title(),
            a.get("account_type", ""),
            a.get("institution", ""),
            a.get("account_last4", ""),
            _money(a.get("balance")),
            _money(a.get("cash_balance")),
            _money(a.get("floor_amount")),
            a.get("source_notes", ""),
        ])
    _table(ws, 3, ["Source", "Owner", "Category", "Type", "Institution", "Last 4", "Balance", "Cash", "Floor", "Source Notes"], rows, money_cols={7, 8, 9})


def _trust_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "Trust Assets", 1, 1, 5)
    rows = [[t.get("source", "Profile"), t.get("description", ""), t.get("property_address", ""), _money(t.get("value")), t.get("source_notes", "")] for t in snapshot.get("trust_assets", [])]
    _table(ws, 3, ["Source", "Description", "Property Address", "Value", "Source Notes"], rows, money_cols={4})


def _liabilities_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "Liabilities", 1, 1, 7)
    rows = []
    for l in snapshot.get("liabilities", []):
        rows.append([l.get("source", "Profile"), l.get("liability_type", ""), l.get("lender", ""), l.get("interest_rate", ""), l.get("account_last4", ""), _money(l.get("balance")), l.get("source_notes", "")])
    _table(ws, 3, ["Source", "Type", "Lender", "Rate", "Last 4", "Balance", "Source Notes"], rows, money_cols={6})


def _inputs_sheet(ws, snapshot: dict[str, Any]) -> None:
    _title(ws, "Raw Inputs", 1, 1, 3)
    rows = [[key, value] for key, value in sorted(snapshot.get("inputs", {}).items())]
    _table(ws, 3, ["Field", "Value"], rows)


def _title(ws, text: str, row: int, col: int, span: int) -> None:
    ws.cell(row, col, text)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
    cell = ws.cell(row, col)
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="left")


def _section(ws, text: str, row: int, col: int) -> None:
    ws.cell(row, col, text)
    ws.cell(row, col).font = Font(bold=True, color=INK)
    ws.cell(row, col).fill = PatternFill("solid", fgColor=BLUE_SOFT)


def _key_values(ws, start_row: int, rows: list[tuple[str, Any]], *, money: bool = False) -> None:
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = Font(bold=True, color=MUTED)
        cell = ws.cell(row, 2)
        if money:
            cell.value = _money(value)
            cell.number_format = MONEY_FMT
        else:
            cell.value = value


def _table(ws, start_row: int, headers: list[str], rows: list[list[Any]], *, money_cols: set[int] | None = None) -> None:
    money_cols = money_cols or set()
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=GRAY)
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            if col_idx in money_cols:
                cell.number_format = MONEY_FMT
    if not rows:
        ws.cell(start_row + 1, 1, "No records")


def _finish_sheet(ws) -> None:
    thin = Side(style="thin", color="D9E1EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = max(len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
    ws.freeze_panes = "A3"


def _money(value: Any) -> float:
    return float(parse_money(value))


def _owner(value: Any) -> str:
    return {"0": "Joint", "1": "Client 1", "2": "Client 2"}.get(str(value), f"Client {value}")
