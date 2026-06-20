from __future__ import annotations

import json
import os
import sqlite3
import tempfile
import unittest
from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from aw_portal.calculations import age_on, calculate_report
from aw_portal.db import init_db
from aw_portal.excel import build_excel_workbook
from aw_portal.pdfs import build_sacs_pdf, build_tcc_pdf
from aw_portal.security import can, hash_password, mask_ssn_last4, verify_password
from aw_portal.web import ensure_client_access, ensure_user_access, load_client_bundle, persist_report_only_profile_items, save_client, save_user, snapshot_from_form, store_report
from seed_demo import seed_demo


def sample_snapshot() -> dict:
    return {
        "report_date": "2026-06-18",
        "client": {"id": 1, "household_name": "Sample Household"},
        "people": [
            {"person_index": 1, "first_name": "Alex", "last_name": "Walker", "dob": "1980-06-01", "ssn_last4": "1234"},
            {"person_index": 2, "first_name": "Jordan", "last_name": "Walker", "dob": "1982-07-15", "ssn_last4": "5678"},
        ],
        "inputs": {
            "report_date": "2026-06-18",
            "inflow": "15000",
            "outflow": "11000",
            "insurance_deductibles": "3500",
            "private_reserve_balance": "42000",
            "investment_account_balance": "180000",
        },
        "accounts": [
            {"id": 1, "owner_index": 1, "category": "retirement", "account_type": "IRA", "institution": "Schwab", "account_last4": "1111", "label": "Schwab IRA", "balance": "11000", "cash_balance": "1000"},
            {"id": 2, "owner_index": 1, "category": "retirement", "account_type": "Roth IRA", "institution": "Schwab", "account_last4": "2222", "label": "Schwab Roth IRA", "balance": "15000", "cash_balance": "500"},
            {"id": 3, "owner_index": 2, "category": "retirement", "account_type": "401K", "institution": "Schwab", "account_last4": "3333", "label": "Schwab 401K", "balance": "50000", "cash_balance": "0"},
            {"id": 4, "owner_index": 0, "category": "non_retirement", "account_type": "Brokerage", "institution": "Schwab", "account_last4": "4444", "label": "Joint Brokerage", "balance": "70000", "cash_balance": "2500"},
        ],
        "trust_assets": [
            {"id": 1, "description": "Primary Residence", "property_address": "100 Peachtree St, Atlanta, GA", "value": "450000"}
        ],
        "liabilities": [
            {"id": 1, "liability_type": "Mortgage", "interest_rate": "6.25", "account_last4": "9999", "balance": "200000"}
        ],
        "deductible_items": [
            {"id": 1, "label": "Home", "amount": "2500"},
            {"id": 2, "label": "Auto", "amount": "1000"},
        ],
    }


class CalculationTests(unittest.TestCase):
    def test_age_on_handles_birthdays(self) -> None:
        self.assertEqual(age_on("1980-06-18", date(2026, 6, 18)), 46)
        self.assertEqual(age_on("1980-06-19", date(2026, 6, 18)), 45)

    def test_report_calculations_match_prd_rules(self) -> None:
        snapshot = calculate_report(sample_snapshot())
        totals = snapshot["totals"]
        self.assertEqual(totals["excess"], "4000.00")
        self.assertEqual(totals["private_reserve_target"], "69500.00")
        self.assertEqual(totals["client_1_retirement"], "26000.00")
        self.assertEqual(totals["client_2_retirement"], "50000.00")
        self.assertEqual(totals["non_retirement"], "70000.00")
        self.assertEqual(totals["trust"], "450000.00")
        self.assertEqual(totals["grand_total"], "596000.00")
        self.assertEqual(totals["liabilities"], "200000.00")

    def test_missing_required_balance_fails(self) -> None:
        snapshot = sample_snapshot()
        snapshot["accounts"][0]["balance"] = ""
        with self.assertRaises(ValueError):
            calculate_report(snapshot)

    def test_private_reserve_override_wins_when_set(self) -> None:
        snapshot = sample_snapshot()
        snapshot["client"]["private_reserve_target_override"] = "80000"
        self.assertEqual(calculate_report(snapshot)["totals"]["private_reserve_target"], "80000.00")


class PdfTests(unittest.TestCase):
    def test_sacs_and_tcc_pdfs_generate(self) -> None:
        snapshot = calculate_report(sample_snapshot())
        sacs = build_sacs_pdf(snapshot)
        tcc = build_tcc_pdf(snapshot)
        self.assertTrue(sacs.startswith(b"%PDF"))
        self.assertTrue(tcc.startswith(b"%PDF"))
        self.assertGreater(len(sacs), 1000)
        self.assertGreater(len(tcc), 1000)


class ExcelTests(unittest.TestCase):
    def test_excel_workbook_generates_expected_sheets_and_values(self) -> None:
        snapshot = calculate_report(sample_snapshot())
        body = build_excel_workbook(snapshot, {"id": 7, "created_at": "2026-06-18 12:00:00", "status": "generated", "export_status": "not_configured"})
        self.assertGreater(len(body), 5000)
        workbook = load_workbook(BytesIO(body), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "SACS", "TCC Totals", "Accounts", "Trust Assets", "Liabilities", "Inputs"],
        )
        self.assertEqual(workbook["Summary"]["B3"].value, "Sample Household")
        self.assertEqual(workbook["Summary"]["B20"].value, 596000)
        self.assertEqual(workbook["Accounts"]["G5"].number_format, '"$"#,##0')
        self.assertEqual(workbook["Liabilities"]["F4"].value, 200000)


class PersistenceTests(unittest.TestCase):
    def test_report_snapshot_survives_profile_change(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        client_id = save_client(
            conn,
            {
                "household_name": "Walker Household",
                "monthly_salary": "15000",
                "monthly_expense_budget": "11000",
                "insurance_deductibles": "3500",
                "p1_first_name": "Alex",
                "p1_last_name": "Walker",
                "p1_dob": "1980-06-01",
                "p1_ssn_last4": "1234",
                "account_1_category": "retirement",
                "account_1_owner_index": "1",
                "account_1_account_type": "IRA",
                "account_1_institution": "Schwab",
                "account_1_account_last4": "1111",
                "trust_1_description": "Primary Residence",
                "trust_1_property_address": "100 Peachtree St",
                "liability_1_liability_type": "Mortgage",
                "liability_1_interest_rate": "6.25",
            },
        )
        bundle = load_client_bundle(conn, client_id)
        account_id = bundle["accounts"][0]["id"]
        trust_id = bundle["trust_assets"][0]["id"]
        liability_id = bundle["liabilities"][0]["id"]
        snapshot = snapshot_from_form(
            bundle,
            {
                "report_date": "2026-06-18",
                "inflow": "15000",
                "outflow": "11000",
                "insurance_deductibles": "3500",
                "private_reserve_balance": "42000",
                "investment_account_balance": "180000",
                f"account_{account_id}_balance": "26000",
                f"account_{account_id}_cash_balance": "1000",
                f"trust_{trust_id}_value": "450000",
                f"liability_{liability_id}_balance": "200000",
            },
        )
        calculate_report(snapshot)
        report_id = store_report(conn, client_id, snapshot)
        files = [row["file_type"] for row in conn.execute("SELECT file_type FROM generated_files WHERE report_id = ? ORDER BY file_type", (report_id,))]
        self.assertEqual(files, ["excel", "sacs", "tcc"])
        save_client(
            conn,
            {
                "household_name": "Changed Household",
                "monthly_salary": "1",
                "monthly_expense_budget": "1",
                "p1_first_name": "Changed",
                "p1_last_name": "Name",
            },
            client_id,
        )
        stored = conn.execute("SELECT snapshot_json FROM quarterly_reports WHERE id = ?", (report_id,)).fetchone()
        stored_snapshot = json.loads(stored["snapshot_json"])
        self.assertEqual(stored_snapshot["client"]["household_name"], "Walker Household")
        self.assertEqual(stored_snapshot["totals"]["grand_total"], "476000.00")

    def test_deductible_items_feed_report_snapshot(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        client_id = save_client(
            conn,
            {
                "household_name": "Deductible Household",
                "monthly_salary": "12000",
                "monthly_expense_budget": "8000",
                "p1_first_name": "Dana",
                "p1_last_name": "Client",
                "deductible_1_label": "Home",
                "deductible_1_amount": "2500",
                "deductible_2_label": "Auto",
                "deductible_2_amount": "1000",
                "account_1_category": "non_retirement",
                "account_1_owner_index": "1",
                "account_1_account_type": "Brokerage",
            },
        )
        bundle = load_client_bundle(conn, client_id)
        account_id = bundle["accounts"][0]["id"]
        snapshot = snapshot_from_form(
            bundle,
            {
                "report_date": "2026-06-18",
                "inflow": "12000",
                "outflow": "8000",
                "private_reserve_balance": "10000",
                f"account_{account_id}_balance": "50000",
            },
        )
        calculate_report(snapshot)
        self.assertEqual(snapshot["inputs"]["insurance_deductibles"], "3500.00")
        self.assertEqual(snapshot["totals"]["private_reserve_target"], "51500.00")

    def test_report_only_assets_are_snapshotted_and_optional_profile_items(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        client_id = save_client(
            conn,
            {
                "household_name": "Report Only Household",
                "monthly_salary": "10000",
                "monthly_expense_budget": "6000",
                "p1_first_name": "Riley",
                "p1_last_name": "Client",
                "account_1_category": "retirement",
                "account_1_owner_index": "1",
                "account_1_account_type": "IRA",
            },
        )
        bundle = load_client_bundle(conn, client_id)
        account_id = bundle["accounts"][0]["id"]
        snapshot = snapshot_from_form(
            bundle,
            {
                "report_date": "2026-06-18",
                "inflow": "10000",
                "outflow": "6000",
                "private_reserve_balance": "22000",
                f"account_{account_id}_balance": "50000",
                "ro_account_1_category": "non_retirement",
                "ro_account_1_owner_index": "0",
                "ro_account_1_account_type": "Report Brokerage",
                "ro_account_1_institution": "Manual",
                "ro_account_1_balance": "25000",
                "ro_account_1_add_to_profile": "1",
                "ro_trust_1_description": "Report Property",
                "ro_trust_1_property_address": "1 Main St",
                "ro_trust_1_value": "300000",
                "ro_liability_1_liability_type": "Report Loan",
                "ro_liability_1_balance": "90000",
            },
        )
        calculate_report(snapshot)
        self.assertEqual(snapshot["totals"]["non_retirement"], "25000.00")
        self.assertEqual(snapshot["totals"]["trust"], "300000.00")
        self.assertEqual(snapshot["totals"]["liabilities"], "90000.00")
        self.assertEqual(snapshot["totals"]["grand_total"], "375000.00")
        persist_report_only_profile_items(conn, client_id, snapshot)
        reloaded = load_client_bundle(conn, client_id)
        self.assertEqual(len(reloaded["accounts"]), 2)
        self.assertEqual(reloaded["accounts"][1]["account_type"], "Report Brokerage")
        self.assertEqual(len(reloaded["trust_assets"]), 0)
        self.assertEqual(len(reloaded["liabilities"]), 0)


class SecurityTests(unittest.TestCase):
    def test_password_hashing_and_roles(self) -> None:
        encoded = hash_password("correct horse")
        self.assertTrue(verify_password("correct horse", encoded))
        self.assertFalse(verify_password("wrong", encoded))
        self.assertTrue(can("system_admin", "company_admin"))
        self.assertTrue(can("company_admin", "planner"))
        self.assertTrue(can("planner", "assistant"))
        self.assertFalse(can("viewer", "assistant"))

    def test_ssn_masking(self) -> None:
        self.assertEqual(mask_ssn_last4("1234"), "***-****")
        self.assertEqual(mask_ssn_last4("1234", reveal=True), "***-1234")

    def test_user_save_hashes_password(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        user_id = save_user(
            conn,
            {
                "email": "planner@example.com",
                "full_name": "Planner User",
                "role": "planner",
                "is_active": "1",
                "password": "Secret123!",
            },
        )
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        self.assertEqual(row["role"], "planner")
        self.assertTrue(verify_password("Secret123!", row["password_hash"]))

    def test_company_admin_cannot_create_admin_or_cross_company_user(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        company_admin = {"id": 1, "role": "company_admin", "company_id": 1}
        with self.assertRaises(PermissionError):
            save_user(
                conn,
                {
                    "email": "sys@example.com",
                    "full_name": "Sys Admin",
                    "role": "system_admin",
                    "company_id": "1",
                    "password": "Secret123!",
                },
                actor=company_admin,
            )
        user_id = save_user(
            conn,
            {
                "email": "viewer@example.com",
                "full_name": "Viewer User",
                "role": "viewer",
                "company_id": "999",
                "password": "Secret123!",
            },
            actor=company_admin,
        )
        row = conn.execute("SELECT company_id, role FROM users WHERE id = ?", (user_id,)).fetchone()
        self.assertEqual(row["company_id"], 1)
        self.assertEqual(row["role"], "viewer")

    def test_tenant_access_blocks_other_company_records(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        init_db(conn)
        other_company = conn.execute("INSERT INTO companies (name, slug) VALUES (?, ?)", ("Other Co", "other-co")).lastrowid
        client_id = save_client(
            conn,
            {
                "household_name": "Other Household",
                "monthly_salary": "10000",
                "monthly_expense_budget": "7000",
                "p1_first_name": "Other",
                "p1_last_name": "Client",
            },
            company_id=other_company,
        )
        with self.assertRaises(PermissionError):
            ensure_client_access(conn, {"role": "planner", "company_id": 1}, client_id)
        ensure_client_access(conn, {"role": "system_admin", "company_id": 1}, client_id)

    def test_company_admin_cannot_manage_admin_users(self) -> None:
        with self.assertRaises(PermissionError):
            ensure_user_access(
                {"role": "company_admin", "company_id": 1},
                {"role": "company_admin", "company_id": 1},
            )


class SeedTests(unittest.TestCase):
    def test_demo_seed_is_idempotent(self) -> None:
        previous = os.environ.get("RAILWAY_DATABASE_PATH")
        with tempfile.TemporaryDirectory() as temp_dir:
            os.environ["RAILWAY_DATABASE_PATH"] = os.path.join(temp_dir, "demo.sqlite3")
            first = seed_demo()
            second = seed_demo()
        if previous is None:
            os.environ.pop("RAILWAY_DATABASE_PATH", None)
        else:
            os.environ["RAILWAY_DATABASE_PATH"] = previous
        self.assertGreaterEqual(first["users_created"], 5)
        self.assertGreaterEqual(first["reports_created"], 4)
        self.assertEqual(second["users_created"], 0)
        self.assertEqual(second["reports_created"], 0)
        self.assertEqual(second["drafts_created"], 0)


if __name__ == "__main__":
    unittest.main()
